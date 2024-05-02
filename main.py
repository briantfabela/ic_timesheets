import os
import openpyxl
from datetime import datetime, timedelta

def generate_refined_timesheets(start_date, end_date):
    timesheets = []
    current_date = start_date

    while current_date < end_date:
        end_date_of_period = current_date + timedelta(days=13)
        if end_date_of_period.month != current_date.month:
            last_day_of_month = current_date.replace(day=28) + timedelta(days=4)
            last_day_of_month = last_day_of_month - timedelta(days=last_day_of_month.day)
            timesheets.append((current_date, last_day_of_month, ' (p1)'))
            start_of_next_month = last_day_of_month + timedelta(days=1)
            end_of_next_period = start_of_next_month + timedelta(days=(end_date_of_period - last_day_of_month).days - 1)
            timesheets.append((start_of_next_month, end_of_next_period, ' (p2)'))
            current_date = end_of_next_period + timedelta(days=1)
        else:
            timesheets.append((current_date, end_date_of_period, ''))
            current_date = end_date_of_period + timedelta(days=1)
        if current_date >= end_date:
            break

    return timesheets

def create_timestamped_folder(base_path):
    """
    Create a timestamped folder within the given base path to store the files.
    Returns the path of the newly created folder.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    new_folder_path = os.path.join(base_path, f'Timesheets_{timestamp}')
    os.makedirs(new_folder_path, exist_ok=True)
    return new_folder_path

def get_column_letter(index):
    return openpyxl.utils.cell.get_column_letter(index)

def clear_cell_range(sheet, start_col, end_col, start_row=4, end_row=4):
    """Clears the specified range of cells."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            sheet.cell(row=row, column=col).value = None

def create_and_save_timesheets(timesheets, base_directory, template_path, sheet_name='Timesheet'):
    file_paths = []
    start_column_index = 4  # Column D is the 4th column
    end_column_index = 17  # Column Q is the 17th column

    for index, (start_date, end_date, part) in enumerate(timesheets):
        wb = openpyxl.load_workbook(template_path)
        sheet = wb[sheet_name]
        sheet['S1'] = end_date.strftime('%-m/%-d/%y')

        if part == ' (p1)':
            cell_column_start = start_column_index
            cell_column_end = start_column_index + (end_date - start_date).days
            # Clear unused columns in p1
            clear_cell_range(sheet, cell_column_end + 1, end_column_index, 6, 11)
        elif part == ' (p2)':
            previous_end_date = timesheets[index - 1][1]
            days_in_p1 = (previous_end_date - timesheets[index - 1][0]).days + 1
            cell_column_start = start_column_index + days_in_p1
            cell_column_end = end_column_index
            # Clear unused columns in p2
            clear_cell_range(sheet, start_column_index, cell_column_start - 1, 6, 11)
        else:
            cell_column_start = start_column_index
            cell_column_end = end_column_index

        # Clear day values in row 4 from D4 to Q4 before filling
        clear_cell_range(sheet, start_column_index, end_column_index, 4, 4)

        # Fill the day values in row 4 within the defined range
        current_date = start_date
        cell_column = cell_column_start
        while current_date <= end_date and cell_column <= cell_column_end:
            sheet.cell(row=4, column=cell_column).value = current_date.day
            current_date += timedelta(days=1)
            cell_column += 1

        # Save the workbook
        filename = f"BF Timesheet {end_date.strftime('%y-%m-%d')}{part}.xlsx"
        file_path = os.path.join(base_directory, filename)
        wb.save(file_path)
        file_paths.append(file_path)

    return file_paths

# Define the time range for the timesheets
start_date = datetime(2023, 6, 16)
end_date = datetime(2024, 5, 1)

# Generate the timesheet periods
timesheets = generate_refined_timesheets(start_date, end_date)

# Set the directory path for Google Drive
directory_path = '/content/drive/MyDrive/Timesheets'
os.makedirs(directory_path, exist_ok=True)

# Path to your Excel template - adjust the filename as needed
template_path = '/content/drive/MyDrive/Timesheets/BF Timesheet Template.xlsx'

timesheet_folder_path = create_timestamped_folder(directory_path)

# Create and save the timesheet files
created_files = create_and_save_timesheets(timesheets, timesheet_folder_path, template_path)

# Optionally print the paths of the created files
for file_path in created_files:
    print(file_path)
